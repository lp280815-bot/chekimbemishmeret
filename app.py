import io
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def match_checks(bank_bytes: bytes, matching_bytes: bytes) -> bytes:
    bank_wb = load_workbook(io.BytesIO(bank_bytes))
    matching_wb = load_workbook(io.BytesIO(matching_bytes))

    bank_ws = bank_wb.active
    matching_ws = matching_wb.active

    def headers_map(ws):
        return {cell.value: cell.column for cell in ws[1] if cell.value}

    bank_headers = headers_map(bank_ws)
    matching_headers = headers_map(matching_ws)

    bank_amount_col = bank_headers.get("סכום השיק בשח")
    bank_ref_col = bank_headers.get("אסמכתא")

    matching_amount_col = matching_headers.get("סכום")
    matching_ref2_col = matching_headers.get("אסמכתא 2")
    matching_ref1_col = matching_headers.get("אסמכתא 1")
    matching_match_col = matching_headers.get("מס. התאמה")

    found_map = {}

    # כלל 1 – סכום + אסמכתא 2
    for row in range(2, bank_ws.max_row + 1):
        bank_amount = bank_ws.cell(row=row, column=bank_amount_col).value
        bank_ref_val = bank_ws.cell(row=row, column=bank_ref_col).value
        bank_ref = str(bank_ref_val).strip() if bank_ref_val is not None else ""

        found = False

        if bank_amount is None or bank_ref == "":
            found_map[row] = 0
            continue

        for m_row in range(2, matching_ws.max_row + 1):
            m_amount = matching_ws.cell(row=m_row, column=matching_amount_col).value
            m_ref2_val = matching_ws.cell(row=m_row, column=matching_ref2_col).value
            m_ref2 = str(m_ref2_val).strip() if m_ref2_val is not None else ""
            m_match = matching_ws.cell(row=m_row, column=matching_match_col).value

            if m_match != 0:
                continue

            if m_amount == bank_amount and m_ref2 == bank_ref:
                matching_ws.cell(row=m_row, column=matching_match_col).value = 1
                found = True
                break

        found_map[row] = 1 if found else 0

    result_col = bank_ws.max_column + 1
    bank_ws.cell(row=1, column=result_col).value = "נמצא במשטח התאמות"

    for row in range(2, bank_ws.max_row + 1):
        bank_ws.cell(row=row, column=result_col).value = found_map.get(row, 0)

    # כלל 2 – סכום + אסמכתא 1
    for row in range(2, bank_ws.max_row + 1):
        if bank_ws.cell(row=row, column=result_col).value == 1:
            continue

        bank_amount = bank_ws.cell(row=row, column=bank_amount_col).value
        bank_ref_val = bank_ws.cell(row=row, column=bank_ref_col).value
        bank_ref = str(bank_ref_val).strip() if bank_ref_val is not None else ""

        if bank_amount is None or bank_ref == "":
            continue

        for m_row in range(2, matching_ws.max_row + 1):
            m_amount = matching_ws.cell(row=m_row, column=matching_amount_col).value
            m_ref1_val = matching_ws.cell(row=m_row, column=matching_ref1_col).value
            m_ref1 = str(m_ref1_val).strip() if m_ref1_val is not None else ""
            m_match = matching_ws.cell(row=m_row, column=matching_match_col).value

            if m_match != 0:
                continue

            if m_amount == bank_amount and m_ref1 == bank_ref:
                matching_ws.cell(row=m_row, column=matching_match_col).value = 1
                bank_ws.cell(row=row, column=result_col).value = 1
                break

    bank_ws.sheet_view.rightToLeft = True
    for row in bank_ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal="right")

    result_wb = matching_wb

    if "פירוט שקים מהבנק" in result_wb.sheetnames:
        result_wb.remove(result_wb["פירוט שקים מהבנק"])

    result_wb.create_sheet("פירוט שקים מהבנק")
    bank_copy_ws = result_wb["פירוט שקים מהבנק"]

    for row in bank_ws.iter_rows(values_only=True):
        bank_copy_ws.append(row)

    output = io.BytesIO()
    result_wb.save(output)
    output.seek(0)
    return output.getvalue()


st.title("התאמת שקים מהבנק מול משטח התאמות")

bank_file = st.file_uploader("העלי פירוט שקים מהבנק", type=["xlsx"])
matching_file = st.file_uploader("העלי משטח עבודה לניתוח כרטיסים", type=["xlsx"])

if bank_file and matching_file:
    if st.button("הרץ התאמה"):
        result_bytes = match_checks(bank_file.read(), matching_file.read())

        st.download_button(
            label="הורדת קובץ תוצאה",
            data=result_bytes,
            file_name="תוצאת_התאמת_שקים.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
